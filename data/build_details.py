#!/usr/bin/env python3
"""区詳細ページ用の追加データ集計。出力: data/processed/ward-details.json"""
import csv, json
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / 'raw'
OUT = Path(__file__).parent / 'processed' / 'ward-details.json'
WARD_IDS = [f'131{i:02d}' for i in range(1, 24)]


def land_price():
    """地価公示 住宅地（用途0）の区別平均（円/㎡）"""
    result = {}
    with open(RAW / 'chika_r7_chiten.csv', encoding='cp932') as f:
        rows = list(csv.reader(f))
    header = rows[1]
    ci = {name: header.index(name) for name in ['都道府県市区町村コード', '標準地番号（用途）', '当年価格（円）']}
    acc = {}
    for row in rows[2:]:
        if len(row) <= max(ci.values()):
            continue
        code = row[ci['都道府県市区町村コード']].strip()
        youto = row[ci['標準地番号（用途）']].strip()
        price = row[ci['当年価格（円）']].replace(',', '').strip()
        if code in WARD_IDS and youto == '0' and price.isdigit():
            acc.setdefault(code, []).append(int(price))
    for code, prices in acc.items():
        result[code] = {'land_price_avg': round(sum(prices) / len(prices)), 'land_price_points': len(prices)}
    return result


def _total_population():
    """住民基本台帳「世帯と人口」から区別総人口（年少+生産年齢+老年）を算出。"""
    pop = {}
    with open(RAW / 'jy26qv0301.csv', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    header = rows[0]
    code_i = header.index('地域コード')
    total_cols = [i for i, name in enumerate(header) if name.endswith('／総数(人)')]
    for row in rows[1:]:
        if len(row) <= code_i:
            continue
        code = row[code_i].strip()
        if code not in WARD_IDS:
            continue
        try:
            total = sum(int(row[i].replace(',', '')) for i in total_cols)
        except (ValueError, IndexError):
            continue
        pop[code] = total
    return pop


def foreign_rate():
    """区別 外国人人口比率（％） = 外国人人口 ÷ 総人口(住基台帳) × 100

    データソース: 東京都「国籍・地域別　外国人人口」（東京都の統計）
    data/raw/ga26ev0100.csv （区市町村別・国籍別 外国人人口, 令和8年1月1日現在）
    分母は data/raw/jy26qv0301.csv （住民基本台帳 世帯と人口, 令和8年1月1日現在）の
    年少+生産年齢+老年人口の合計。
    """
    src = RAW / 'ga26ev0100.csv'
    if not src.exists():
        return {}
    pop = _total_population()
    if not pop:
        return {}
    foreign = {}
    with open(src, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    header = rows[0]
    try:
        code_i = header.index('地域コード')
        foreign_i = header.index('総数')
    except ValueError:
        return {}
    for row in rows[1:]:
        if len(row) <= max(code_i, foreign_i):
            continue
        code = row[code_i].strip()
        if code not in WARD_IDS:
            continue
        val = row[foreign_i].replace(',', '').strip()
        if val.isdigit():
            foreign[code] = int(val)
    result = {}
    for code in WARD_IDS:
        if code in foreign and code in pop and pop[code] > 0:
            result[code] = {'foreign_rate': round(foreign[code] / pop[code] * 100, 2)}
    return result


def income():
    """納税義務者1人当たり課税対象所得（千円）。

    総務省「市町村税課税状況等の調」（令和6年度）〔市町村別内訳〕第11表
    「課税標準額段階別令和６年度分所得割額等に関する調（合計）
    （所得割納税義務者数・課税対象所得・課税標準額・所得割額）」
    （data/raw/soumu_J51-24-b.xlsx）の市町村民税行から、
        課税対象所得（千円） ÷ 所得割の納税義務者数
    で1人当たり課税対象所得を算出する。

    出典URL: https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/xls/J51-24-b.xlsx
    （総務省「令和6年度 市町村税課税状況等の調」 ichiran09_24.html 第11表の市町村別内訳）
    列: 1=団体コード（6桁・検査数字付き）, 4=表側（市町村民税/道府県民税）,
        5=所得割の納税義務者数（人）, 13=課税対象所得（千円）
    """
    src = RAW / 'soumu_J51-24-b.xlsx'
    if not src.exists():
        return {}
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result = {}
    for row in ws.iter_rows(values_only=True):
        code6 = str(row[1]).strip() if row[1] else ''
        code = code6[:5]
        if code not in WARD_IDS or row[4] != '市町村民税':
            continue
        taxpayers, taxable_income = row[5], row[13]
        if isinstance(taxpayers, (int, float)) and isinstance(taxable_income, (int, float)) and taxpayers > 0:
            result[code] = {'income_per_taxpayer': round(taxable_income / taxpayers)}
    return result


def crime(population):
    """人口千人当たり刑法犯認知件数。

    警視庁「区市町村の町丁別、罪種別及び手口別認知件数（年累計）」令和6年分
    （data/raw/keishicho_R6_ninchikensu.csv, CC BY 4.0, Shift_JIS）から、
    区名そのものが「市区町丁」列の値と完全一致する行（区内町丁の小計行）の
    「総合計」列を採用する。

    このCSVは町丁ごとの明細行に加え、区・市ごとの合計行（市区町丁列が
    区名のみの行）を別途含んでいる。区名前方一致で明細行を合算すると
    区名のみの合計行まで二重に加算してしまうため、区名との完全一致行の
    値をそのまま使う（合算しない）。
    """
    src = RAW / 'keishicho_R6_ninchikensu.csv'
    if not src.exists():
        return {}
    names = {}
    with open(Path(__file__).parent / 'processed' / 'wards.json', encoding='utf-8') as f:
        for w in json.load(f)['wards']:
            names[w['id']] = w['name']
    with open(src, encoding='cp932') as f:
        rows = list(csv.reader(f))
    header = rows[0]
    total_i = header.index('総合計')
    by_name = {}
    for row in rows[1:]:
        if len(row) <= total_i:
            continue
        by_name[row[0].strip()] = row[total_i].strip()
    result = {}
    for code in WARD_IDS:
        name = names.get(code)
        val = by_name.get(name)
        if name and val and val.isdigit() and population.get(code):
            result[code] = {'crime_per_1000': round(int(val) / population[code] * 1000, 1)}
    return result


def waiting_children():
    """待機児童数（人）。

    こども家庭庁「保育所等関連状況取りまとめ（令和7年4月1日）」の
    （参考）資料1～6（data/raw/cfa_hoiku_r7_shiryo.xlsx）から、
    資料6-1「待機児童数が減少した市区町村の状況」と
    資料6-2「待機児童数が増加（変化なしを含む）した市区町村の状況」の
    東京都行を読み、「R7.4」列（令和7年4月1日現在）の値を採用する。
    両シートとも1行に2組の表（都道府県・市区町村・R7.4・R6.4）が並ぶため、
    列オフセット2起点と8起点の両方を走査する。

    資料6は「待機児童数が令和6年及び令和7年ともにゼロの市区町村は除く」
    仕様であり、23区のうち未掲載の区は待機児童数ゼロとして扱う。この
    「未掲載＝ゼロ」の解釈が崩れると全区が無言でゼロになるため、資料4
    「全国待機児童マップ（市区町村別）」の東京都計と、資料6から拾った
    東京都全市区町村の合計が一致することを検証してから区へ割り当てる。

    出典: こども家庭庁「保育所等関連状況取りまとめ（令和7年4月1日）」
    https://www.cfa.go.jp/policies/hoiku/torimatome/r7
    利用条件: 公共データ利用規約（PDL1.0）。出典と加工の明示を条件に
    商用・非商用を問わず二次利用できる（こども家庭庁コピーライトポリシー
    https://www.cfa.go.jp/copyright-policy ）。
    区名から区コードへは Task 7 と同じく data/processed/wards.json の names を逆引きする。
    """
    src = RAW / 'cfa_hoiku_r7_shiryo.xlsx'
    if not src.exists():
        return {}
    names = {}
    with open(Path(__file__).parent / 'processed' / 'wards.json', encoding='utf-8') as f:
        for w in json.load(f)['wards']:
            names[w['name']] = w['id']
    wb = openpyxl.load_workbook(src, data_only=True)

    # 資料6の東京都行（区部・市部・町村部すべて）を拾う
    tokyo = {}
    for sheet in ('資料６－１', '資料６－２'):
        for row in wb[sheet].iter_rows(values_only=True):
            for pref_i, city_i, r7_i in ((2, 3, 4), (8, 9, 10)):
                if len(row) <= r7_i:
                    continue
                if row[pref_i] != '東京都' or not isinstance(row[city_i], str):
                    continue
                if isinstance(row[r7_i], (int, float)):
                    tokyo[row[city_i].strip()] = int(row[r7_i])
    if not tokyo:
        return {}

    # 資料4の東京都計と突合し、「未掲載＝ゼロ」の前提を検証する
    expected = None
    for row in wb['資料4'].iter_rows(values_only=True):
        if len(row) > 16 and row[15] == '東京都' and isinstance(row[16], (int, float)):
            expected = int(row[16])
            break
    actual = sum(tokyo.values())
    assert expected is not None, 'waiting_children: 資料4に東京都の待機児童数計が見つからない'
    assert actual == expected, (
        f'waiting_children: 資料6の東京都合計 {actual} が資料4の東京都計 {expected} と一致しない。'
        '資料の様式変更が疑われるため、未掲載区をゼロとみなす前提を再確認すること'
    )

    # 資料6に載らない区は待機児童ゼロ
    return {code: {'waiting_children': tokyo.get(name, 0)} for name, code in names.items()}


def top_stations():
    """区ごとの乗降人員上位3駅。

    調査したが「区マッピング付き・23区完備」のクリーンなデータ源が
    見つからなかったため、意図的に未実装（常に空を返す）。呼び出し側は
    これを missing 扱いにしてゲートで指標ごと落とす。

    調査済みで採用しなかった候補:
    - 国土数値情報 S12 駅別乗降客数（全国, GML/Shapefile）:
      駅の所在区カラムがなく、空間結合が必要だが geopandas/pyshp 等の
      GIS ライブラリが環境に無い。
    - 東京都交通局 各駅乗降人員一覧（都営地下鉄4路線, CSV）:
      駅名のみで区カラムなし。かつ都営地下鉄は世田谷・杉並・足立・
      葛飾・江戸川区などを通らず23区を網羅できない。
    - 東京都統計年鑑 4-8/4-13/4-15（JR/私鉄/地下鉄, 2019年データ）:
      事業者ごとに別ファイルで区カラムがなく、手動の駅→区マッピングが
      必要な上にデータが2019年時点で古い。
    """
    return {}


def main():
    lp = land_price()
    missing = [w for w in WARD_IDS if w not in lp]
    assert not missing, f'land_price missing wards: {missing}'  # kill test: 23区揃うこと

    sources = {'land_price': '国土交通省 令和7年地価公示（住宅地 区別平均・円/㎡）'}

    pop = _total_population()
    pop_missing = [w for w in WARD_IDS if w not in pop]
    assert not pop_missing, f'population missing wards: {pop_missing}'  # kill test: 23区揃うこと
    sources['population'] = '住民基本台帳による世帯と人口（令和8年1月1日現在）・人'

    cr = crime(pop)
    cr_missing = [w for w in WARD_IDS if w not in cr]
    assert not cr_missing, f'crime missing wards: {cr_missing}'  # kill test: 23区揃うこと
    sources['crime'] = '警視庁「区市町村の町丁別、罪種別及び手口別認知件数」（令和6年分・人口千人当たり）'

    wc = waiting_children()
    wc_missing = [w for w in WARD_IDS if w not in wc]
    assert not wc_missing, f'waiting_children missing wards: {wc_missing}'  # kill test: 23区揃うこと
    sources['waiting_children'] = 'こども家庭庁「保育所等関連状況取りまとめ」（令和7年4月1日現在）資料6・人'

    inc = income()
    inc_missing = [w for w in WARD_IDS if w not in inc]
    if inc_missing:
        print(f'income DROPPED: missing wards {inc_missing or "(no source file)"}')
        inc = {}
    else:
        sources['income'] = '総務省「市町村税課税状況等の調」（令和6年度）第11表 課税対象所得 ÷ 所得割の納税義務者数・千円'

    fr = foreign_rate()
    fr_missing = [w for w in WARD_IDS if w not in fr]
    if fr_missing:
        print(f'foreign_rate DROPPED: missing wards {fr_missing or "(no source file)"}')
        fr = {}
    else:
        sources['foreign_rate'] = '東京都の統計「国籍・地域別 外国人人口」（令和8年1月1日現在）÷ 住民基本台帳による世帯と人口 総人口（令和8年1月1日現在）・％'

    ts = top_stations()
    ts_missing = [w for w in WARD_IDS if w not in ts]
    if ts_missing:
        print(f'top_stations DROPPED: missing wards {ts_missing or "(no source data)"}')
        ts = {}
    else:
        sources['stations'] = '国土数値情報 S12 駅別乗降客数'

    wards = []
    for w in WARD_IDS:
        entry = {'id': w, **lp[w], 'population': pop[w], **cr[w], **wc[w]}
        if w in inc:
            entry.update(inc[w])
        if w in fr:
            entry.update(fr[w])
        if w in ts:
            entry['top_stations'] = ts[w]
        wards.append(entry)

    data = {'sources': sources, 'wards': wards}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding='utf-8')
    print(f'wrote {OUT} ({len(wards)} wards)')


if __name__ == '__main__':
    main()
