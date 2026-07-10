#!/usr/bin/env python3
"""23区×5軸データセット生成スクリプト

生データ (data/raw/) から各軸の指標を抽出し、
data/processed/wards.json / wards.csv を生成する。

軸と出典:
  賑わい軸: 昼夜間人口比率      — 令和2年国勢調査 (tokyo_daytime_population_2020_tj20zv0100.csv)
  成熟軸:   高齢化率・年少人口率 — 住民基本台帳 年齢別人口 (jy26qv0301.csv)
  みどり軸: 一人当たり公立公園面積 — 東京都建設局 公園調書 令和7年4月 (kouen_r7_kushichoson.csv)
  世帯軸:   単身世帯率・子育て世帯率 — 令和2年国勢調査 家族類型 (estat_table10_kazokutypes.xlsx)
  華やか軸: 財政力指数          — 総務省 主要財政指標 令和5年度 (soumu_000983094.xlsx)
"""
import csv
import json
import os
import re

RAW = os.path.join(os.path.dirname(__file__), "raw")
OUT = os.path.join(os.path.dirname(__file__), "processed")

WARDS = [
    "千代田区", "中央区", "港区", "新宿区", "文京区", "台東区", "墨田区",
    "江東区", "品川区", "目黒区", "大田区", "世田谷区", "渋谷区", "中野区",
    "杉並区", "豊島区", "北区", "荒川区", "板橋区", "練馬区", "足立区",
    "葛飾区", "江戸川区",
]


def num(s):
    if s is None:
        return None
    s = str(s).replace(",", "").replace("％", "").strip()
    if s in ("", "-", "－"):
        return None
    return float(s)


def load_daytime():
    """昼夜間人口比率（総数, %）"""
    out = {}
    path = os.path.join(RAW, "tokyo_daytime_population_2020_tj20zv0100.csv")
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            name = row["地域"].strip()
            if name in WARDS:
                out[name] = num(row["昼夜間人口比率／総数（％）"])
    return out


def load_age():
    """高齢化率・年少人口率（住民基本台帳, %）"""
    out = {}
    path = os.path.join(RAW, "jy26qv0301.csv")
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            name = row[2].strip()
            if name in WARDS:
                young = num(row[3])
                working = num(row[6])
                old = num(row[9])
                total = young + working + old
                out[name] = {
                    "aging_rate": round(old / total * 100, 2),
                    "youth_rate": round(young / total * 100, 2),
                }
    return out


def load_park():
    """一人当たり公立公園面積（ハ/Ｂ, ㎡）— 国民公園(皇居等)・公団は除外"""
    out = {}
    path = os.path.join(RAW, "kouen_r7_kushichoson.csv")
    with open(path, encoding="cp932") as f:
        for row in csv.reader(f):
            if len(row) < 26:
                continue
            name = row[1].strip()
            if name in WARDS:
                out[name] = num(row[25])  # 公立公園合計 一人当たり面積 ハ/Ｂ
    return out


def load_household():
    """単身世帯率・子育て世帯率（令和2年国勢調査 第10表, %）"""
    import openpyxl

    out = {}
    path = os.path.join(RAW, "estat_table10_kazokutypes.xlsx")
    ws = openpyxl.load_workbook(path, read_only=True).active
    totals = {}   # ward -> (総世帯数, 子育て世帯数)
    singles = {}  # ward -> 単身世帯数
    for row in ws.iter_rows(min_row=10, values_only=True):
        area = str(row[2] or "")
        m = re.match(r"131\d\d_(.+)", area)
        if not m or m.group(1) not in WARDS:
            continue
        name = m.group(1)
        size = str(row[3] or "")
        if size.startswith("0_"):  # 総数行
            total = num(row[4])
            kids = (num(row[8]) or 0) + (num(row[9]) or 0) + (num(row[10]) or 0)
            totals[name] = (total, kids)
        elif size.startswith("1_"):  # 世帯人員1人
            singles[name] = num(row[4])
    for name, (total, kids) in totals.items():
        out[name] = {
            "single_rate": round(singles[name] / total * 100, 2),
            "family_rate": round(kids / total * 100, 2),
        }
    return out


def load_fiscal():
    """財政力指数（令和5年度）"""
    import openpyxl

    out = {}
    path = os.path.join(RAW, "soumu_000983094.xlsx")
    ws = openpyxl.load_workbook(path, read_only=True).active
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] == "東京都" and row[2] in WARDS:
            out[row[2]] = num(row[3])
    return out


def main():
    os.makedirs(OUT, exist_ok=True)
    daytime = load_daytime()
    age = load_age()
    park = load_park()
    household = load_household()
    fiscal = load_fiscal()

    # カバレッジ検証: 全23区が全指標で揃うこと
    for label, d in [("昼夜間人口比率", daytime), ("年齢構成", age),
                     ("公園面積", park), ("世帯構成", household),
                     ("財政力指数", fiscal)]:
        missing = [w for w in WARDS if w not in d or d[w] is None]
        assert not missing, f"{label}: 欠損 {missing}"
        print(f"OK {label}: 23/23区")

    wards = []
    for i, name in enumerate(WARDS):
        wards.append({
            "id": f"131{i + 1:02d}",
            "name": name,
            "metrics": {
                "daytime_population_ratio": daytime[name],
                "aging_rate": age[name]["aging_rate"],
                "youth_rate": age[name]["youth_rate"],
                "park_area_per_capita": park[name],
                "single_household_rate": household[name]["single_rate"],
                "family_household_rate": household[name]["family_rate"],
                "fiscal_strength_index": fiscal[name],
            },
        })

    meta = {
        "generated_from": "data/raw/",
        "sources": {
            "daytime_population_ratio": "令和2年国勢調査 従業地・通学地による人口・就業状態等集計（東京都, CC BY 4.0）",
            "aging_rate/youth_rate": "住民基本台帳による東京都の世帯と人口 年齢別人口（令和8年1月）",
            "park_area_per_capita": "東京都建設局 都市公園等区市町村別面積・人口割合表（令和7年4月1日現在）公立公園合計ハ/Ｂ",
            "single/family_household_rate": "令和2年国勢調査 人口等基本集計 第10表 世帯の家族類型",
            "fiscal_strength_index": "総務省 全市町村の主要財政指標（令和5年度）",
        },
        "wards": wards,
    }

    with open(os.path.join(OUT, "wards.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    keys = ["daytime_population_ratio", "aging_rate", "youth_rate",
            "park_area_per_capita", "single_household_rate",
            "family_household_rate", "fiscal_strength_index"]
    with open(os.path.join(OUT, "wards.csv"), "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "name"] + keys)
        for ward in wards:
            w.writerow([ward["id"], ward["name"]] + [ward["metrics"][k] for k in keys])

    print(f"\nwrote {OUT}/wards.json, wards.csv ({len(wards)} wards)")


if __name__ == "__main__":
    main()
